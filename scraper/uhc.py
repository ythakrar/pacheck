import os, sys, logging, requests, psycopg2
from datetime import datetime
from io import BytesIO
try:
    import openpyxl
except ImportError:
    print("pip install openpyxl requests psycopg2-binary")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://localhost/pacheck")

UHC_SOURCES = [
    {"plan_type": "commercial", "url": "https://www.uhcprovider.com/content/dam/provider/docs/public/prior-auth/comm-prior-auth-notification-list.xlsx"},
    {"plan_type": "medicare",   "url": "https://www.uhcprovider.com/content/dam/provider/docs/public/prior-auth/ma-prior-auth-notification-list.xlsx"},
    {"plan_type": "medicaid",   "url": "https://www.uhcprovider.com/content/dam/provider/docs/public/prior-auth/community-plan-prior-auth-list.xlsx"},
]

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; PACheck-Bot/1.0)"}

def get_db():
    return psycopg2.connect(DATABASE_URL)

def get_payer_id(cur, slug="uhc"):
    cur.execute("SELECT id FROM payers WHERE slug = %s", (slug,))
    row = cur.fetchone()
    if not row:
        raise ValueError(f"Payer '{slug}' not found.")
    return row[0]

def get_plan_type_id(cur, slug):
    cur.execute("SELECT id FROM plan_types WHERE slug = %s", (slug,))
    row = cur.fetchone()
    return row[0] if row else None

def get_or_create_cpt(cur, code, description="", specialty="", category=""):
    code = code.strip().upper()
    cur.execute("SELECT id FROM cpt_codes WHERE code = %s", (code,))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(
        "INSERT INTO cpt_codes (code, description, specialty, category) VALUES (%s,%s,%s,%s) RETURNING id",
        (code, description[:500] if description else "", specialty, category),
    )
    return cur.fetchone()[0]

def upsert_pa_rule(cur, cpt_id, payer_id, plan_type_id, status, notes, turnaround, portal, source_url):
    cur.execute(
        "SELECT id, status FROM pa_rules WHERE cpt_id=%s AND payer_id=%s AND plan_type_id=%s AND state IS NULL",
        (cpt_id, payer_id, plan_type_id),
    )
    existing = cur.fetchone()
    if not existing:
        cur.execute(
            """INSERT INTO pa_rules
               (cpt_id, payer_id, plan_type_id, status, notes, turnaround_days,
                submission_portal, source_url, source_type, confidence, scraped_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'scraper',4,%s)""",
            (cpt_id, payer_id, plan_type_id, status, notes, turnaround, portal, source_url, datetime.utcnow()),
        )
        return "inserted"
    rule_id, old_status = existing
    if old_status != status:
        cur.execute(
            "UPDATE pa_rules SET status=%s, notes=%s, turnaround_days=%s, scraped_at=%s WHERE id=%s",
            (status, notes, turnaround, datetime.utcnow(), rule_id),
        )
        return "updated"
    cur.execute("UPDATE pa_rules SET scraped_at=%s WHERE id=%s", (datetime.utcnow(), rule_id))
    return "unchanged"

def categorize_code(code, desc, specialty):
    d = (desc + " " + specialty).lower()
    if any(w in d for w in ["mri","ct scan","x-ray","radiol","imaging","ultrasound","echo"]):
        return "radiology"
    if any(w in d for w in ["therapy","psycho","behav","mental","counseling","psychiatric"]):
        return "behavioral_health"
    if any(w in d for w in ["cardiac","cardio","heart","coronary"]):
        return "cardiology"
    if any(w in d for w in ["orthop","arthroplasty","knee","hip","shoulder","spine","joint"]):
        return "orthopedic"
    if any(w in d for w in ["oncol","chemo","cancer","tumor","infusion"]):
        return "oncology"
    if any(w in d for w in ["endoscopy","colonoscopy","gastro"]):
        return "gastroenterology"
    return "other"

def build_default_notes(status, plan_type, payer_name):
    if status == "required":
        return f"Prior authorization required per {payer_name} published PA list. Submit clinical documentation and medical necessity prior to service."
    if status == "conditional":
        return f"PA may be required depending on specific plan and clinical indication. Verify with {payer_name} before service."
    return f"No prior authorization required per {payer_name} published guidelines."

def parse_uhc_excel(workbook, plan_type, source_url):
    results = []
    for sheet in workbook.worksheets:
        col_map = {}
        header_row = 1
        for row_idx, row in enumerate(sheet.iter_rows(max_row=15, values_only=True), 1):
            cells = [str(c).lower().strip() if c else "" for c in row]
            if any("cpt" in c or "hcpcs" in c or c == "code" for c in cells):
                header_row = row_idx
                for i, c in enumerate(cells):
                    if "cpt" in c or "hcpcs" in c or c == "code": col_map["code"] = i
                    elif "desc" in c:                              col_map["desc"] = i
                    elif any(x in c for x in ["auth","prior","required","pa"]):
                                                                   col_map["status"] = i
                    elif "spec" in c or "categ" in c:             col_map["specialty"] = i
                    elif "note" in c or "criteria" in c:          col_map["notes"] = i
                break
        if "code" not in col_map:
            continue
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            raw_code = row[col_map["code"]] if row else None
            if not raw_code:
                continue
            code = str(raw_code).strip().upper().replace(" ","")
            if len(code) < 4 or len(code) > 7:
                continue
            desc     = str(row[col_map["desc"]])      if "desc"     in col_map and row[col_map["desc"]]     else ""
            specialty= str(row[col_map["specialty"]])  if "specialty"in col_map and row[col_map["specialty"]] else ""
            notes    = str(row[col_map["notes"]])      if "notes"    in col_map and row[col_map["notes"]]     else ""
            raw_status = str(row[col_map["status"]]).lower() if "status" in col_map and row[col_map["status"]] else "yes"
            if any(x in raw_status for x in ["yes","required","y"]):
                status, turnaround = "required", "3-5 days"
            elif any(x in raw_status for x in ["no","not required","n","exempt"]):
                status, turnaround = "not_required", "N/A"
            elif any(x in raw_status for x in ["conditional","varies","cond"]):
                status, turnaround = "conditional", "3-5 days"
            else:
                status, turnaround = "required", "3-5 days"
            category = categorize_code(code, desc, specialty)
            results.append({
                "code": code, "desc": desc.strip()[:500],
                "specialty": specialty.strip()[:128], "category": category,
                "status": status, "turnaround": turnaround,
                "notes": notes.strip()[:2000] or build_default_notes(status, plan_type, "UHC"),
                "portal": "Optum PA Portal", "source_url": source_url,
            })
    return results

def scrape_uhc():
    log.info("UHC scraper starting")
    conn = get_db()
    cur  = conn.cursor()
    payer_id = get_payer_id(cur, "uhc")
    stats = {"found":0,"inserted":0,"updated":0,"unchanged":0,"errors":0}
    cur.execute("INSERT INTO scraper_runs (payer_id, status) VALUES (%s,'running') RETURNING id", (payer_id,))
    run_id = cur.fetchone()[0]
    conn.commit()
    for source in UHC_SOURCES:
        plan_type    = source["plan_type"]
        url          = source["url"]
        plan_type_id = get_plan_type_id(cur, plan_type)
        if not plan_type_id:
            continue
        log.info(f"Fetching {plan_type}: {url}")
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            workbook = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
        except Exception as e:
            log.error(f"Download failed: {e}")
            stats["errors"] += 1
            continue
        records = parse_uhc_excel(workbook, plan_type, url)
        stats["found"] += len(records)
        for rec in records:
            try:
                cpt_id = get_or_create_cpt(cur, rec["code"], rec["desc"], rec["specialty"], rec["category"])
                result = upsert_pa_rule(cur, cpt_id, payer_id, plan_type_id, rec["status"],
                                        rec["notes"], rec["turnaround"], rec["portal"], rec["source_url"])
                stats[result] = stats.get(result, 0) + 1
            except Exception as e:
                log.error(f"Error saving {rec.get('code')}: {e}")
                stats["errors"] += 1
        conn.commit()
    cur.execute(
        "UPDATE scraper_runs SET status=%s, completed_at=%s, codes_found=%s, codes_added=%s, codes_updated=%s WHERE id=%s",
        ("success" if not stats["errors"] else "partial", datetime.utcnow(),
         stats["found"], stats["inserted"], stats["updated"], run_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    log.info(f"UHC done: {stats}")
    return stats

if __name__ == "__main__":
    scrape_uhc()
